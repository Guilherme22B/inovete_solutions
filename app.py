estoque = {}

def adicionar_item(item, quantidade):
    if item in estoque:
        estoque[item] += quantidade
    else:
        estoque[item] = quantidade

def remover_item(item, quantidade):
    if item in estoque:
        estoque[item] -= quantidade
        if estoque[item] <= 0:
            del estoque[item]
        return True
    else:
        return False

def exibir_estoque():
    print("Estoque:")
    for item, quantidade in estoque.items():
        print(f"{item}: {quantidade}")

while True:
    print("1 - Adicionar item")
    print("2 - Remover item")
    print("3 - Exibir estoque")
    print("4 - Calcular escala de vendas")
    print("5 - Calcular imposto de renda")
    print("6 - Sair")
    opcao = input("Escolha uma opção: ")

    if opcao == "1":
        item = input("Digite o nome do item: ")
        quantidade = int(input("Digite a quantidade: "))
        adicionar_item(item, quantidade)
        print("Item adicionado ao estoque.")
    elif opcao == "2":
        item = input("Digite o nome do item: ")
        quantidade = int(input("Digite a quantidade: "))
        if remover_item(item, quantidade):
            print("Item removido do estoque.")
        else:
            print("Item não encontrado no estoque.")
    elif opcao == "3":
        exibir_estoque()
    elif opcao == "4":
        margem_lucro = float(input("Digite a margem de lucro: " ))
        preco_custo = float(input("Digite o preço de custo do produto: " ))
        quantidade = int(input("Digite a quantidade de produtos vendidos: " ))

        margem_lucro = margem_lucro + 1 

        calcular_vendas = margem_lucro * preco_custo * quantidade

        preco_venda = preco_custo * margem_lucro

        lucro_produto = preco_venda - preco_custo

        total_vendas = preco_venda * quantidade

        escala_vendas = calcular_vendas

        print("O lucro por produto é de R$", lucro_produto)
        print("A escala de vendas é de R$", escala_vendas)

    elif opcao == "5":
        lucrototal = float(input("Informe o lucro total obtido com as vendas: "))

        impostorenda = 0

        if lucrototal <= 18000:
            impostorenda = lucrototal * 0.1

        elif lucrototal <= 45000:
            imposto1 = 18000 * 0.1
            imposto2 = (lucrototal - 18000) * 0.15
            impostorenda = imposto1 + imposto2

        elif lucrototal <= 98000:
            imposto1 = 18000 * 0.1
            imposto2 = (45000 - 18000) * 0.15
            imposto3 = (lucrototal - 45000) * 0.20
            impostorenda = imposto1 + imposto2 + imposto3

        else:
            imposto1 = 18000 * 0.1
            imposto2 = (45000 - 18000) * 0.15
            imposto3 = (98000 - 45000) * 0.20
            imposto4 = (lucrototal - 98000) * 0.25
            impostorenda = imposto1 + imposto2 + imposto3 + imposto4

        valor_liquido = lucrototal - impostorenda

        print("Imposto de Renda: R$", impostorenda)
        print("Valor líquido: R$", valor_liquido)

    elif opcao == "6":
        break

    else:
        print("Opção inválida. Digite novamente.")

    
#conversor para exel

from openpyxl import Workbook

arquivo  = Workbook()

Projeto = arquivo.active #criação da planilha

plan_qte_estoque = arquivo.create_sheet("qte_estoque") #criação da subi planilha
plan_itens = arquivo.create_sheet("Itens_(lista)") #criação da subi planilha
plan_escala_venda = arquivo.create_sheet("escala de venda") #criação da subi planilha
plan_imposto_renda = arquivo.create_sheet("imposto de renda") #criação da subi planilha
plan_mag_lucro = arquivo.create_sheet("margem de lucro") #criação da subi planilha
plan_custo = arquivo.create_sheet("custo") #criação da subi planilha
plan_preco_venda = arquivo.create_sheet("preço de venda") #criação da subi planilha
plan_total_vendas = arquivo.create_sheet("total de vendas") #criação da subi planilha
plan_lucro_total = arquivo.create_sheet("lucro total") #criação da subi planilha
Projeto.title = "Produtos"

a = str(input("qualquer coisa")) #teste para colocar informações
b = str(input("qualquer coisa")) #teste para colocar informações
c = str(input("qualquer coisa")) #teste para colocar informações 
d = str(input("qualquer coisa")) #teste para colocar informações

Projeto["A1"] = a #A1 igual a celula que vai ficar com os dados A2 A3 A4 a mesma coisa
Projeto["A2"] = b
Projeto["A3"] = c
Projeto["A4"] = d

plan_itens["A1"] = a #plan_itens igual a planilha que vai ser salva os dados 
plan_itens["A2"] = b
plan_itens["A3"] = c
plan_itens["A4"] = d

arquivo.save("Projeto.xlsx") #teste para salvar as planilhas